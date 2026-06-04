"""Data Export Utilities -- Rotation Task B (Part 3)

Export sentiment, keyword and topic analysis results to CSV and Excel formats.
Supports per-subreddit breakdowns, raw thread lists, aggregated stats, and
full LDA/KeyBERT outputs as separate sheets or CSV rows.

Usage:
    from visual.export_util import SentimentExporter, KeywordExporter, TopicExporter, BatchExporter

    # Export sentiment analysis results
    exporter = SentimentExporter(stats_data)
    paths = exporter.to_files("/tmp/sentiment_report")   # writes .csv & .xlsx

    # Or combine multiple analyses in one workbook
    batch = BatchExporter()
    batch.add("sentiment", sentiment_result)
    batch.add("keywords",  keybert_result)
    batch.add("topics",    lda_result)
    xlsx_path = batch.to_excel("/tmp/full_analysis.xlsx")

Dependencies: pip install pandas openpyxl (Excel format only)
"""

from __future__ import annotations

import csv
import io
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any, Optional


# --------------------------------------------------------------------------- #
#  Helpers                                                                    #
# --------------------------------------------------------------------------- #

def _flatten_row(d: dict | object) -> dict:
    """Flatten nested dicts/lists into dot-separated keys so each row has flat columns."""
    out: dict[str, Any] = {}
    for k, v in (vars(d).items() if is_dataclass(d) else d.items()):
        key = str(k)
        if isinstance(v, (dict, object)) and not isinstance(v, (str, int, float, bool)):
            try:
                sub = asdict(v) if is_dataclass(v) else _try_dict(v)
                for sk2, sv2 in sub.items():
                    out[f"{key}.{sk2}"] = sv2
            except Exception:
                out[key] = str(v)
        elif isinstance(v, (list, tuple)):
            if len(v) <= 5 and all(not isinstance(item, (dict, list)) for item in v):
                # Short lists — store semi-colon separated string
                out[key] = ";".join(str(x) for x in v)
            else:
                out[key] = str(v[:20])  # truncate long sequences
        else:
            out[key] = v if v is not None else ""
    return out


def _try_dict(o: Any) -> dict | None:
    try:
        return o.to_dict() if callable(getattr(o, "to_dict", None)) else vars(o)
    except Exception:
        return None


# --------------------------------------------------------------------------- #
#  CSV Exporter                                                               #
# --------------------------------------------------------------------------- #

class CSVExporter:
    """Export analysis results to CSV files.

    Handles dict and dataclass inputs, flattening nested structures where possible.
    """

    def __init__(self) -> None:
        self._rows: list[dict[str, Any]] = []
        self._headers: list[str] | None = None

    # ---- single table -------------------------------------------------------

    def add_rows(self, data: dict | list[dict]) -> "CSVExporter":
        """Append rows (chainable)."""
        if isinstance(data, dict):
            flat = _flatten_row(data)
            self._rows.append(flat)
            if self._headers is None:
                self._headers = list(flat.keys())
        elif isinstance(data, list):
            if not data:
                return self
            first = data[0]
            is_dc = is_dataclass(first)
            flat_items = [asdict(first) if is_dc else _try_dict(first)]
            keys_set: set[str] = set()
            for item in flat_items + [first]:
                for k, v in (vars(item).items() if is_dataclass(item) else (item.items() if isinstance(item, dict) else [])):
                    keys_set.add(str(k))
            self._headers = list(keys_set)
            self._rows.extend([_flatten_row(item) for item in data])
        return self

    def to_csv(self, file_path: str, delimiter: str = ",") -> Optional[str]:
        """Write accumulated rows to a CSV file. Creates parent directories automatically."""
        if not self._rows or not self._headers:
            return None

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=self._headers, delimiter=delimiter, extrasaction="ignore")
            writer.writeheader()
            for row in self._rows:
                writer.writerow({k: row.get(k, "") for k in self._headers})

        return str(path)


# --------------------------------------------------------------------------- #
#  Excel Exporter (multi-sheet)                                               #
# --------------------------------------------------------------------------- #

class ExcelExporter:
    """Export to an Excel workbook with multiple sheets.

    Each analysis result maps to one sheet. Auto-flattens dataclasses and dicts.
    """

    def __init__(self) -> None:
        self._sheets: dict[str, list[dict]] = {}  # name → list of flat rows

    def add_sheet(self, name: str, data: dict | list[dict]) -> "ExcelExporter":
        """Queue a sheet. Chainable."""
        if isinstance(data, dict):
            self._sheets[name] = [_flatten_row(data)]
        elif isinstance(data, list) and data:
            first = data[0]
            is_dc = is_dataclass(first)
            self._sheets[name] = [_flatten_row(item) for item in data]
        return self

    def to_excel(self, file_path: str) -> Optional[str]:
        """Write all queued sheets to an Excel workbook."""
        try:
            import openpyxl     # type: ignore[import-untyped]
        except ImportError:
            raise ImportError("openpyxl is required for Excel export.\n    pip install openpyxl")

        from pandas import DataFrame   # type: ignore[import-untyped]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # default sheet not needed

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        for sheet_name, rows in self._sheets.items():
            if not rows:
                continue
            safe_name = sheet_name[:31]  # xls has 31-char sheet name limit
            ws = wb.create_sheet(title=safe_name)
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])

        wb.save(path)
        return str(path)


# --------------------------------------------------------------------------- #
#  Convenience wrappers                                                       #
# --------------------------------------------------------------------------- #

class SentimentExporter:
    """Export sentiment analysis results to CSV + Excel."""

    def __init__(self, stats_data: dict):
        self.stats_data = stats_data

    def to_files(self, base_path: str = "/tmp/sentiment_export") -> list[str]:
        """Write per-subreddit CSV files plus a summary Excel workbook.

        Returns:
            List of output file paths.
        """
        output_paths: list[str] = []

        # --- CSV: per-subreddit breakdown ----------------------------------
        per_sub = self.stats_data.get("per_subreddit", {})
        if not per_sub:
            return output_paths

        csv_exp = CSVExporter()
        for sub, stats in per_sub.items():
            flat = _flatten_row(stats)
            flat["subreddit"] = sub
            csv_exp.add_rows(flat)

        csv_path = f"{base_path}_per_subreddit.csv"
        if csv_exp.to_csv(csv_path):
            output_paths.append(csv_path)

        # --- Excel: summary + per-thread details ---------------------------
        excel_exp = ExcelExporter()

        # Summary row
        for key in ["total_analyzed", "global_mean_vader", "global_positive_pct",
                     "global_negative_pct", "global_neutral_pct"]:
            if key in self.stats_data:
                excel_exp.add_sheet("summary", {key: self.stats_data[key]})

        per_thread = self.stats_data.get("per_subreddit")  # placeholder access pattern
        # Export the full stats dict as a single-row sheet for reference
        excel_exp.add_sheet("full_stats", _flatten_row(self.stats_data))

        xlsx_path = f"{base_path}.xlsx"
        ep = excel_exp.to_excel(xlsx_path)
        if ep:
            output_paths.append(ep)

        return output_paths


class KeywordExporter:
    """Export keyword / phrase extraction results.

    Handles TFIDFResult, KeyBERTResult, or plain dict with 'per_subreddit' + 'global_themes'.
    """

    def __init__(self, result: Any):
        self.result = result

    def to_csv(self, file_path: str) -> Optional[str]:
        csv_exp = CSVExporter()
        per_sub = {}
        global_kw = []

        # Handle dataclass or dict input
        if hasattr(self.result, "per_subreddit"):  # dataclass with to_dict or asdict attr
            result_dict = self.result.to_dict() if hasattr(self.result, "to_dict") else vars(self.result)
        else:
            result_dict = dict(self.result)

        per_sub = {k: v for k, v in result_dict.get("per_subreddit", {}).items()}
        global_kw = result_dict.get("global_themes", []) or result_dict.get("global_keywords", [])

        for sub, data in per_sub.items():
            themes_list = []
            if isinstance(data, dict):
                for theme_name, theme_data in data.items():
                    if isinstance(theme_data, list):
                        for item in theme_data:
                            flat = _flatten_row(item)
                            flat["subreddit"] = sub
                            flat["theme_name"] = theme_name
                            themes_list.append(flat)
            elif hasattr(data, "themes"):  # SubredditThemes dataclass
                for t in data.themes:
                    flat = _flatten_row(t)
                    flat["subreddit"] = sub
                    themes_list.append(flat)

        if global_kw:
            for kw in global_kw[:50]:
                flat = _flatten_row(kw)
                flat["type"] = "global"
                csv_exp.add_rows(flat)

        # Per-subreddit themes
        for sub, data in per_sub.items():
            for item_data in (data.themes if hasattr(data, "themes") else
                              ([data] if isinstance(data, dict) and "phrase" not in data else [])):
                pass  # Handled above

        if not csv_exp._rows:
            # Fallback: export as simple per-subreddit keywords
            for sub, data in per_sub.items():
                if hasattr(data, "themes"):
                    themes_list = data.themes
                elif isinstance(data, dict) and "themes" in data:
                    themes_list = data["themes"]
                else:
                    continue

                if not isinstance(themes_list, list):
                    continue

                for theme_idx, theme in enumerate(themes_list):
                    flat = _flatten_row(theme) if is_dataclass(theme) or hasattr(theme, "to_dict") \
                        else {k: v for k, v in (theme.items() if isinstance(theme, dict) else [])}  # type: ignore[union-attr]
                    flat["subreddit"] = sub
                    flat["rank_in_sub"] = theme_idx + 1
                    csv_exp.add_rows(flat)

        return csv_exp.to_csv(file_path)


class TopicExporter:
    """Export LDA topic modeling results."""

    def __init__(self, result: Any):
        self.result = result

    def to_csv(self, file_path: str) -> Optional[str]:
        csv_exp = CSVExporter()
        if hasattr(self.result, "to_dict"):
            rdict = self.result.to_dict()   # type: ignore[union-attr]
        elif isinstance(self.result, dict):
            rdict = self.result
        else:
            rdict = vars(self.result)  # type: ignore[arg-type]

        # Topic definitions as rows
        for tid, terms in (rdict.get("topics_per_topic") or {}).items():
            term_strs = []
            for t_idx, (term, weight) in enumerate(terms[:10]):
                flat_line = {f"topic_{tid}.term_{t_idx}": term}
                flat_line[f"topic_{tid}.weight_{t_idx}"] = weight
                csv_exp.add_rows(flat_line)

        # Topic-subreddit breakdown
        bd = rdict.get("topic_subreddit_breakdown") or {}
        for tid, subs in (bd or {}).items():
            sub_data: dict[str, Any] = {"topic_id": tid}
            for sub, avg_prob in subs.items():
                sub_data[f"subreddit:{sub}"] = avg_prob
            csv_exp.add_rows(sub_data)

        # Document-level distributions
        docs = rdict.get("document_distributions") or []
        n_topics = rdict.get("n_topics_modelled", 5)
        for doc in docs[:200]:
            flat = _flatten_row(doc) if is_dataclass(doc) \
                else {k: v for k, v in (doc.items() if isinstance(doc, dict) else [])}  # type: ignore[union-attr]
            csv_exp.add_rows(flat)

        # Model metadata as summary row
        meta_keys = ["n_topics_modelled", "coherence_score", "perplexity", "optimal_n_topics"]
        meta_row: dict[str, Any] = {}
        for k in meta_keys:
            if k in rdict:
                meta_row[k] = rdict[k]  # type: ignore[literal-required]
        if meta_row:
            csv_exp.add_rows(meta_row)

        return csv_exp.to_csv(file_path)


class BatchExporter:
    """Combine multiple analysis results into a single Excel workbook."""

    def __init__(self) -> None:
        self._excel = ExcelExporter()

    def add(self, name: str, data: Any, as_list: bool = True) -> "BatchExporter":
        """Add a named sheet. Set `as_list=False` for single-row dicts."""
        if is_dataclass(data):
            list_ = [data] if hasattr(data, "thread_sentiments") else [data]  # type: ignore[unreachable]
            self._excel.add_sheet(name, list_)
        elif as_list:
            if isinstance(data, (list, tuple)):
                self._excel.add_sheet(name, list(data))
            elif is_dataclass(data):
                self._excel.add_sheet(name, [data])
        else:
            if isinstance(data, dict):
                self._excel.add_sheet(name, data)
        return self

    def to_excel(self, file_path: str) -> Optional[str]:
        return self._excel.to_excel(file_path)


# --------------------------------------------------------------------------- #
#  Demo                                                                       #
# --------------------------------------------------------------------------- #

def run_example() -> None:
    """Generate sample export files from synthetic data."""
    print("=" * 60)
    print("   Data Export Utilities - Demo")
    print("=" * 60)

    # Synthetic sentiment stats
    synthetic_stats = {
        "total_analyzed": 15,
        "global_mean_vader": 0.24,
        "global_positive_pct": 53.33,
        "global_negative_pct": 20.0,
        "global_neutral_pct": 26.67,
        "per_subreddit": {
            "python":     {"mean_vader_compound": 0.4, "positive_count": 8,  "negative_count": 3,  "neutral_count": 2, "n_threads": 13},
            "datascience":{"mean_vader_compound": -0.1, "positive_count": 4,  "negative_count": 5,  "neutral_count": 1, "n_threads": 6},
        },
    }

    csv_exp = CSVExporter()
    for sub, stats in synthetic_stats["per_subreddit"].items():
        flat = {"subreddit": sub}
        flat.update({k: v for k, v in stats.items() if isinstance(v, (int, float))})
        csv_exp.add_rows(flat)

    csv_out = csv_exp.to_csv("/tmp/demo_sentiment_export.csv")
    print(f"\n✓ CSV output: {csv_out}")

    # Print CSV content for demo
    if csv_out and Path(csv_out).exists():
        with open(csv_out) as fh:
            print(fh.read())

    # Synthetic KeyBERT-style result (plain dict)
    synthetic_keywords = {
        "per_subreddit": {
            "python": {
                "themes": [
                    {"phrase": "type hints",          "score": 0.85, "document_count": 5},
                    {"phrase": "asyncio performance", "score": 0.72, "document_count": 3},
                    {"phrase": "dataclasses inheritance", "score": 0.61, "document_count": 2},
                ],
            },
        },
        "global_themes": [
            {"phrase": "machine learning",      "score": 0.95, "document_count": 45},
            {"phrase": "feature engineering",   "score": 0.82, "document_count": 32},
            {"phrase": "transformers fine-tuning", "score": 0.78, "document_count": 28},
        ],
    }

    kw_csv_out = KeywordExporter(synthetic_keywords).to_csv("/tmp/demo_keywords_export.csv")
    print(f"✓ Keywords CSV: {kw_csv_out}")

    # Synthetic LDA-like result (plain dict)
    synthetic_topics = {
        "n_topics_modelled": 3,
        "coherence_score": 0.5421,
        "perplexity": 892.4,
        "optimal_n_topics": 3,
        "topics_per_topic": {
            0: [("machine", 0.35), ("learning", 0.28), ("training", 0.21)],
            1: [("python", 0.31), ("asyncio", 0.24), ("dataclasses", 0.19)],
            2: [("react", 0.29), ("components", 0.23), ("performance", 0.17)],
        },
        "topic_subreddit_breakdown": {
            0: {"machinelearning": 0.45, "datascience": 0.38},
            1: {"python": 0.62, "datascience": 0.15},
            2: {"reactjs": 0.71, "javascript": 0.22},
        },
        "document_distributions": [
            {"doc_id": "t1", "subreddit": "machinelearning", "dominant_topic": 0, "confidence": 0.68},
            {"doc_id": "t2", "subreddit": "python",          "dominant_topic": 1, "confidence": 0.74},
            {"doc_id": "t3", "subreddit": "reactjs",        "dominant_topic": 2, "confidence": 0.59},
        ],
    }

    topic_csv_out = TopicExporter(synthetic_topics).to_csv("/tmp/demo_topics_export.csv")
    print(f"✓ Topics CSV: {topic_csv_out}")

    # Batch export to Excel
    batch = BatchExporter()
    batch.add("sentiment_summary", synthetic_stats, as_list=False)
    batch.add("keywords_global", synthetic_keywords.get("global_themes", []))

    for sub, data in synthetic_keywords.get("per_subreddit", {}).items():
        themes = data.get("themes", []) if isinstance(data, dict) else getattr(data, "themes", []) or []
        batch.add(f"keywords_{sub}", themes, as_list=True)  # type: ignore[arg-type]

    for tid_str, terms in synthetic_topics.get("topics_per_topic", {}).items():
        row_data = {"topic_id": int(tid_str), "terms": ";".join(t[0] for t in terms)}  # type: ignore[union-attr]
        batch.add(f"topic_{tid_str}", {k: v for k, v in [("topic_id", tid_str)]}, as_list=False)  # type: ignore[arg-type]
    batch.add("topic_breakdown", synthetic_topics.get("topic_subreddit_breakdown") or {}, as_list=False)

    xlsx_out = batch.to_excel("/tmp/demo_full_analysis.xlsx")
    print(f"✓ Full workbook (Excel): {xlsx_out}")

    # Verify Excel sheets
    if xlsx_out and Path(xlsx_out).exists():
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_out)
        print(f"\n  Sheet names: {wb.sheetnames}")
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            print(f"    - {ws_name}: {ws.max_row} rows x {ws.max_column} cols")

    print("\n✓ Demo complete. Check /tmp/demo_* files.")


if __name__ == "__main__":
    run_example()
